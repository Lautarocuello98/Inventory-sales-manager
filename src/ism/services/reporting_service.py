from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


class ReportingService:
    def __init__(self, repo):
        self.repo = repo

    def monthly_sales_totals(self, months: int = 6) -> list[tuple[str, float]]:
        return self.repo.monthly_sales_totals(months)

    def cumulative_profit_series(self) -> list[tuple[str, float]]:
        return self.repo.cumulative_profit_series()
    
    def export_sales_report_excel(self, path: str, start_iso: str, end_iso: str) -> None:
        wb = Workbook()

        def money(cell):
            cell.number_format = "#,##0.00"

        def pct(cell):
            cell.number_format = "0.00%"

        def bold_row(ws, r):
            for c in ws[r]:
                c.font = Font(bold=True)

        def set_widths(ws, widths: dict[str, int]):
            for col, w in widths.items():
                ws.column_dimensions[col].width = w

        def add_table(ws, name: str, start_row: int, start_col: int, end_row: int, end_col: int):
            ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
            tab = Table(displayName=name, ref=ref)
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(tab)

        totals, _ = self.repo.sales_summary_between(start_iso, end_iso)
        sales_count, revenue_usd, revenue_ars, profit_usd = totals

        sales_rows = self.repo.list_sales_between(start_iso, end_iso)
        purchases_rows = self.repo.list_purchases_between(start_iso, end_iso)

        spent_usd = sum(float(p.total_usd) for p in purchases_rows) if purchases_rows else 0.0
        net_usd = float(profit_usd) - float(spent_usd)

        # -------- 1) Summary --------
        ws = wb.active
        ws.title = "Summary"
        ws["A1"] = "Summary"
        ws["A1"].font = Font(bold=True, size=14)

        ws["A3"] = "Window"
        ws["B3"] = f"{start_iso}  ->  {end_iso}"

        rows = [
            ("Sales count", int(sales_count), "int"),
            ("Revenue USD", float(revenue_usd), "money"),
            ("Revenue ARS", float(revenue_ars), "money"),
            ("Gross Profit USD", float(profit_usd), "money"),
            ("Purchases/Restock Spent USD", float(spent_usd), "money"),
            ("Net USD (Profit - Spent)", float(net_usd), "money"),
        ]

        start_row = 5
        for i, (label, val, kind) in enumerate(rows):
            r = start_row + i
            ws[f"A{r}"] = label
            ws[f"B{r}"] = val
            if kind == "money":
                money(ws[f"B{r}"])

        set_widths(ws, {"A": 28, "B": 34})

        # -------- 2) Sales Detail --------
        ws2 = wb.create_sheet("Sales Detail")
        ws2.append([
            "Sale ID", "Datetime", "Notes",
            "SKU", "Product Name",
            "Qty", "Unit Price USD", "Unit Cost USD",
            "Line Revenue USD", "Line Profit USD", "Margin %"
        ])
        bold_row(ws2, 1)

        out_row = 2
        for s in sales_rows:
            items = self.repo.sale_items_for_sale(int(s.id))
            for it in items:
                margin_pct = (it.line_margin_usd / it.line_total_usd) if it.line_total_usd else 0.0
                ws2.append([
                    int(s.id), s.datetime, s.notes or "",
                    it.sku, it.name,
                    int(it.qty), float(it.unit_price_usd), float(it.cost_usd),
                    float(it.line_total_usd), float(it.line_margin_usd), float(margin_pct)
                ])
                money(ws2[f"G{out_row}"])
                money(ws2[f"H{out_row}"])
                money(ws2[f"I{out_row}"])
                money(ws2[f"J{out_row}"])
                pct(ws2[f"K{out_row}"])
                out_row += 1

        ws2.freeze_panes = "A2"
        set_widths(ws2, {
            "A": 10, "B": 22, "C": 28,
            "D": 14, "E": 34,
            "F": 6, "G": 16, "H": 16,
            "I": 18, "J": 18, "K": 10
        })
        if ws2.max_row >= 2:
            add_table(ws2, "SalesDetail", 1, 1, ws2.max_row, 11)

        # -------- 3) Purchases --------
        ws3 = wb.create_sheet("Purchases")
        ws3["A1"] = "Purchases / Restock"
        ws3["A1"].font = Font(bold=True, size=14)

        ws3["A3"] = "Total spent USD"
        ws3["B3"] = float(spent_usd)
        money(ws3["B3"])

        ws3.append([])
        ws3.append([
            "Purchase ID", "Datetime", "Vendor", "Notes",
            "SKU", "Product Name",
            "Qty", "Unit Cost USD", "Line Total USD"
        ])
        bold_row(ws3, 5)

        out_row = 6
        for p in purchases_rows:
            items = self.repo.purchase_items_for_purchase(int(p.id))
            for it in items:
                ws3.append([
                    int(p.id), p.datetime, p.vendor or "", p.notes or "",
                    it.sku, it.name,
                    int(it.qty), float(it.unit_cost_usd), float(it.line_total_usd)
                ])
                money(ws3[f"H{out_row}"])
                money(ws3[f"I{out_row}"])
                out_row += 1

        ws3.freeze_panes = "A6"
        set_widths(ws3, {
            "A": 12, "B": 22, "C": 18, "D": 26,
            "E": 14, "F": 34,
            "G": 6, "H": 16, "I": 16
        })
        if ws3.max_row >= 6:
            add_table(ws3, "PurchasesDetail", 5, 1, ws3.max_row, 9)

        wb.save(path)
