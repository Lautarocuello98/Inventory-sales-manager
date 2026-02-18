from __future__ import annotations

from openpyxl import load_workbook

from ism.domain.errors import ValidationError
import logging

log = logging.getLogger(__name__)


class ExcelService:
    def __init__(self, repo, purchase_service, inventory_service):
        self.repo = repo
        self.purchases = purchase_service
        self.inventory = inventory_service

    def import_restock_excel(self, path: str) -> tuple[int, int]:
        """
        Excel represents RESTOCK (delta to add), not absolute stock.
        Headers:
          sku | name | cost_usd | price_usd | stock | min_stock
        """
        wb = load_workbook(path)
        ws = wb.active

        headers = {}
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=col).value
            if isinstance(v, str):
                headers[v.strip().lower()] = col

        required = ["sku", "name", "cost_usd", "price_usd", "stock", "min_stock"]
        for r in required:
            if r not in headers:
                raise ValidationError(f"Missing column header: {r}")

        ok = 0
        skipped = 0

        for row in range(2, ws.max_row + 1):
            try:
                sku = ws.cell(row=row, column=headers["sku"]).value
                name = ws.cell(row=row, column=headers["name"]).value
                cost = ws.cell(row=row, column=headers["cost_usd"]).value
                price = ws.cell(row=row, column=headers["price_usd"]).value
                restock_qty = ws.cell(row=row, column=headers["stock"]).value
                min_stock = ws.cell(row=row, column=headers["min_stock"]).value

                if not sku or not name:
                    skipped += 1
                    continue
                if cost is None or price is None or restock_qty is None or min_stock is None:
                    skipped += 1
                    continue

                sku = str(sku).strip()
                name = str(name).strip()
                cost = float(cost)
                price = float(price)
                restock_qty = int(float(restock_qty))
                min_stock = int(float(min_stock))

                if restock_qty < 0:
                    skipped += 1
                    continue

                existing = self.repo.get_product_by_sku(sku)
                if existing:
                    # Update product fields, keep stock unchanged (stock changes only via purchases)
                    self.repo.upsert_product(sku, name, cost, price, existing.stock, min_stock)

                    if restock_qty > 0:
                        self.purchases.create_purchase(
                            vendor="EXCEL_IMPORT",
                            notes=f"Excel restock (+{restock_qty}) for {sku}",
                            items=[{
                                "product_id": existing.id,
                                "qty": restock_qty,
                                "unit_cost_usd": cost,
                            }]
                        )
                else:
                    # Create product with stock=0, then apply restock as purchase
                    self.repo.upsert_product(sku, name, cost, price, 0, min_stock)
                    created = self.repo.get_product_by_sku(sku)
                    if not created:
                        skipped += 1
                        continue
                    if restock_qty > 0:
                        self.purchases.create_purchase(
                            vendor="EXCEL_IMPORT",
                            notes=f"Initial restock from Excel for {sku}",
                            items=[{
                                "product_id": created.id,
                                "qty": restock_qty,
                                "unit_cost_usd": cost,
                            }]
                        )

                ok += 1
            except Exception as e:
                log.warning("Excel import skipped row %s: %s", row, e)
                skipped += 1

        return ok, skipped
