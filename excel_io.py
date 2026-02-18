from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from db import (
    upsert_product,
    sales_summary_between,
    list_sales_between,
    sale_items_for_sale,
    list_purchases_between,
    purchase_items_for_purchase,
)


# ---------------- IMPORT PRODUCTS ----------------

def import_products_from_excel(path: str) -> tuple[int, int]:
    """
    Headers required:
    sku | name | cost_usd | price_usd | stock | min_stock
    """
    wb = load_workbook(path)
    ws = wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if isinstance(v, str):
            headers[v.strip().lower()] = col

    required = ["sku", "name", "cost_usd", "price_usd", "stock", "min_stock"]
    for r in required:
        if r not in headers:
            raise ValueError(f"Missing column header: {r}")

    ok = 0
    skipped = 0

    for row in range(2, ws.max_row + 1):
        try:
            sku = ws.cell(row=row, column=headers["sku"]).value
            name = ws.cell(row=row, column=headers["name"]).value
            cost = ws.cell(row=row, column=headers["cost_usd"]).value
            price = ws.cell(row=row, column=headers["price_usd"]).value
            stock = ws.cell(row=row, column=headers["stock"]).value
            min_stock = ws.cell(row=row, column=headers["min_stock"]).value

            if not sku or not name:
                skipped += 1
                continue

            upsert_product(
                str(sku).strip(),
                str(name).strip(),
                float(cost),
                float(price),
                int(stock),
                int(min_stock),
            )
            ok += 1
        except Exception:
            skipped += 1

    return ok, skipped


# ---------------- EXPORT REPORT (3 sheets) ----------------

def export_sales_report_excel(path: str, start_iso: str, end_iso: str) -> None:
    wb = Workbook()

    def money(cell):
        cell.number_format = '#,##0.00'

    def pct(cell):
        cell.number_format = '0.00%'

    def bold_row(ws, r):
        for c in ws[r]:
            c.font = Font(bold=True)

    def set_widths(ws, widths: dict[str, int]):
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    def add_table(ws, name: str, start_row: int, start_col: int, end_row: int, end_col: int):
        ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        tab = Table(displayName=name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(tab)

    # ---- Pull DB data ----
    totals, _ = sales_summary_between(start_iso, end_iso)
    sales_count, revenue_usd, revenue_ars, profit_usd = totals

    sales_rows = list_sales_between(start_iso, end_iso)  # (id, datetime, total_usd, fx, total_ars, notes)
    purchases_rows = list_purchases_between(start_iso, end_iso)  # (id, datetime, vendor, total_usd, notes)
    spent_usd = sum(float(p[3]) for p in purchases_rows) if purchases_rows else 0.0
    net_usd = float(profit_usd) - float(spent_usd)

    # ===================== 1) SUMMARY =====================
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "Summary"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Window"
    ws["B3"] = f"{start_iso}  ->  {end_iso}"

    rows = [
        ("Sales count", int(sales_count), "int"),
        ("Revenue USD", float(revenue_usd), "money"),
        ("Revenue ARS", float(revenue_ars), "money"),
        ("Gross Profit USD", float(profit_usd), "money"),
        ("Purchases/Restock Spent USD", float(spent_usd), "money"),
        ("Net USD (Profit - Spent)", float(net_usd), "money"),
    ]

    start_row = 5
    for i, (label, val, kind) in enumerate(rows):
        r = start_row + i
        ws[f"A{r}"] = label
        ws[f"B{r}"] = val
        if kind == "money":
            money(ws[f"B{r}"])

    set_widths(ws, {"A": 28, "B": 34})

    # ===================== 2) SALES DETAIL =====================
    ws2 = wb.create_sheet("Sales Detail")
    ws2.append([
        "Sale ID", "Datetime", "Notes",
        "SKU", "Product Name",
        "Qty", "Unit Price USD", "Unit Cost USD",
        "Line Revenue USD", "Line Profit USD", "Margin %"
    ])
    bold_row(ws2, 1)

    out_row = 2
    for s in sales_rows:
        sale_id, dt, _tot_usd, _fx, _tot_ars, notes = s
        items = sale_items_for_sale(int(sale_id))

        for it in items:
            # (sku, name, qty, unit_price_usd, line_total_usd, cost_usd, line_margin_usd)
            sku, name, qty, unit_price, line_total, cost_usd, line_profit = it

            qty = int(qty)
            unit_price = float(unit_price)
            cost_usd = float(cost_usd)
            line_total = float(line_total)
            line_profit = float(line_profit)
            margin_pct = (line_profit / line_total) if line_total else 0.0

            ws2.append([
                int(sale_id),
                dt,
                notes or "",
                sku,
                name,
                qty,
                unit_price,
                cost_usd,
                line_total,
                line_profit,
                margin_pct
            ])

            money(ws2[f"G{out_row}"])
            money(ws2[f"H{out_row}"])
            money(ws2[f"I{out_row}"])
            money(ws2[f"J{out_row}"])
            pct(ws2[f"K{out_row}"])
            out_row += 1

    ws2.freeze_panes = "A2"
    set_widths(ws2, {
        "A": 10, "B": 22, "C": 28,
        "D": 14, "E": 34,
        "F": 6, "G": 16, "H": 16,
        "I": 18, "J": 18, "K": 10
    })
    if ws2.max_row >= 2:
        add_table(ws2, "SalesDetail", 1, 1, ws2.max_row, 11)

    # ===================== 3) PURCHASES =====================
    ws3 = wb.create_sheet("Purchases")
    ws3["A1"] = "Purchases / Restock"
    ws3["A1"].font = Font(bold=True, size=14)

    ws3["A3"] = "Total spent USD"
    ws3["B3"] = float(spent_usd)
    money(ws3["B3"])

    ws3.append([])

    header_row = 5
    ws3.append([
        "Purchase ID", "Datetime", "Vendor", "Notes",
        "SKU", "Product Name",
        "Qty", "Unit Cost USD", "Line Total USD"
    ])
    bold_row(ws3, header_row)

    out_row = header_row + 1
    for p in purchases_rows:
        purchase_id, dt, vendor, total_usd, notes = p
        items = purchase_items_for_purchase(int(purchase_id))

        for it in items:
            # (sku, name, qty, unit_cost_usd, line_total_usd)
            sku, name, qty, unit_cost, line_total = it

            qty = int(qty)
            unit_cost = float(unit_cost)
            line_total = float(line_total)

            ws3.append([
                int(purchase_id),
                dt,
                vendor or "",
                notes or "",
                sku,
                name,
                qty,
                unit_cost,
                line_total
            ])

            money(ws3[f"H{out_row}"])
            money(ws3[f"I{out_row}"])
            out_row += 1

    ws3.freeze_panes = "A6"
    set_widths(ws3, {
        "A": 12, "B": 22, "C": 18, "D": 26,
        "E": 14, "F": 34,
        "G": 6, "H": 16, "I": 16
    })
    if ws3.max_row >= 6:
        add_table(ws3, "PurchasesDetail", 5, 1, ws3.max_row, 9)

    wb.save(path)
